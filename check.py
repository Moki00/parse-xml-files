import requests
import pandas as pd
import lxml.etree as ETREE
import glob
import os
import logging
from typing import Dict, List, Any, Optional
from datetime import datetime
from openpyxl.utils import column_index_from_string
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from dotenv import load_dotenv

# Basic logging to print messages to the console
logging.basicConfig(
    level=logging.INFO, 
    format='%(asctime)s - %(levelname)s - %(message)s'
)

load_dotenv()  # Load variables from .env file

class TeamDynamixSandboxClient:
    """A client for interacting with the TeamDynamix Sandbox Web API."""

    def __init__(self, base_url: str, asset_app_id: int):
        """Initializes the client with the base URL and asset app ID."""
        self.base_url = base_url
        self.asset_app_id = asset_app_id
        self.session = requests.Session()
        self.session.headers.update({"Content-Type": "application/json"})
    
    def authenticate(self, username, password) -> bool:
        """Authenticates with the API and stores the token in the session headers."""
        AUTH_URL = f"{self.base_url}/api/auth/login"
        logging.info("Attempting to authenticate...")

        if not username or not password:
            logging.error("Username or password environment variables are not set.")
            return False

        try:
            response = self.session.post(AUTH_URL, json={"username": username, "password": password})
            response.raise_for_status()  # Raises HTTPError for bad responses (4xx or 5xx)
            
            bearer_token = response.text
            # Set the auth header for all subsequent requests in this session
            self.session.headers.update({"Authorization": f"Bearer {bearer_token}"})
            
            logging.info("✅ Authentication successful!")
            return True
            
        except requests.exceptions.HTTPError as e:
            logging.error(f"Authentication failed: {e.response.status_code} {e.response.reason}")
            if e.response.status_code == 401:
                logging.error("-> Please check that the username and password are correct.")
            return False
        except Exception as e:
            logging.error(f"An unexpected error occurred during authentication: {e}")
            return False
        
    def get_all_assets(self, form_id: int) -> Optional[List[Dict[str, Any]]]:
        """Fetches a list of all assets that use a specific form."""
        search_url = f"{self.base_url}/api/{self.asset_app_id}/assets/search"
        search_body = {"FormIds": [form_id]}
        
        logging.info(f"Fetching all assets with form ID {form_id}...")
        try:
            response = self.session.post(search_url, json=search_body)
            response.raise_for_status()
            
            # The response is a JSON list of asset objects
            assets = response.json()
            logging.info(f"✅ Retrieved {len(assets)} assets.")
            return assets
            
        except requests.exceptions.HTTPError as e:
            logging.error(f"Failed to fetch assets: {e.response.status_code} {e.response.reason}")
            return None
        except Exception as e:
            logging.error(f"An unexpected error occurred while fetching assets: {e}")
            return None
    
    def get_asset_details(self, asset_id: int) -> Optional[Dict[str, Any]]:
        """Fetches detailed information for a specific asset by its ID."""
        detail_url = f"{self.base_url}/api/{self.asset_app_id}/assets/{asset_id}"
        
        logging.info(f"Fetching details for asset ID {asset_id}...")
        try:
            response = self.session.get(detail_url)
            response.raise_for_status()
            
            # The response is a JSON object with detailed asset information
            asset_details = response.json()
            logging.info(f"✅ Retrieved details for asset ID {asset_id}.")
            return asset_details
            
        except requests.exceptions.HTTPError as e:
            logging.error(f"Failed to fetch details for asset ID {asset_id}: {e.response.status_code} {e.response.reason}")
            return None
        except Exception as e:
            logging.error(f"An unexpected error occurred while fetching details for asset ID {asset_id}: {e}")
            return None

CHECKS_TO_PERFORM = [

    # ----------------------------------------------------------
    # ------ TDMA = Phase 2 ------
    # ----------------------------------------------------------

    # -- Phase 2 Voice Capable --
    {
        'group_name': 'Phase 2 Voice Capable',
        'base_xpath': ".//Recset[@Name='Trunking System']/Node[contains(@ReferenceKey, 'GWINNETT')]/Section[@Name='ASTRO 25']",
        'context_node_name': 'Trunking System',
        'fields': {
            'Phase 2 Voice Capable': 'True'
        }
    },

    # -- TDMA Channel ID 3 --
    {
        'group_name': 'Trunking System - Channel ID 3',
        'base_xpath': ".//Recset[@Name='Trunking System']//EmbeddedNode[@ReferenceKey='Channel ID 3']",
        'context_node_name': 'Trunking System',
        'fields': {
            'Identifier Enable': 'True',
            'Base Frequency (MHz)': '851.012500',
            'Channel Spacing (kHz)': '12.500',
            'Channel Type': 'TDMA',
            'Transmit Offset (MHz)': '45.000000',
            'Transmit Offset Sign': '-'
        }
    },

    # -- TDMA Channel ID 4 --
    {
        'group_name': 'Trunking System - Channel ID 4',
        'base_xpath': ".//Recset[@Name='Trunking System']//EmbeddedNode[@ReferenceKey='Channel ID 4']",
        'context_node_name': 'Trunking System',
        'fields': {
            'Identifier Enable': 'True',
            'Base Frequency (MHz)': '762.006250',
            'Channel Spacing (kHz)': '12.500',
            'Channel Type': 'TDMA',
            'Transmit Offset (MHz)': '30.000000',
            'Transmit Offset Sign': '+'
        }
    },

    # ----------------------------------------------------------
    # ------ Personalities ------
    # ----------------------------------------------------------

    # -- 8CALL90 Personality
    {
        'group_name': '8CALL90 Personality',
        'base_xpath': ".//Recset[@Name='Conventional Personality']//EmbeddedNode[@ReferenceKey='8CALL90']",
        'context_node_name': 'Conventional Personality',
        'fields': {
            'Rx / TA Frequency (MHz)': '851.012500',
            'User Selectable PL (MPL)':'False',
            'Tx Squelch Type':'PL',
            'Tx DPL Code':'023',
            'Tx DPL Invert':'False',
            'Rx / TA Squelch Type':'PL',
            'Tx Frequency (MHz)':'806.012500',
            'Tx Network ID':'659',
            'Tx PL Code':'5A',
            'Tx PL Freq':'156.7',
            'Rx / TA  PL Code':'5A',
            'Rx / TA PL Freq':'156.7',
            'Rx / TA DPL Code':'023',
            'Rx / TA DPL Invert':'False',
            'Rx / TA  Network ID':'659',
            'Direct / Talkaround':'False',
            'Direct Squelch Type':'PL',
            'Direct PL Freq':'67.0',
            'Direct PL Code':'XZ',
            'Tx Deviation / Channel Spacing': '4 kHz / 20 kHz',
            'Name':'8CALL90',
            'Direct Network ID':'659',
            'User Selectable PL [MPL]':'Disabled',
            'Direct Frequency (MHz)':'851.012500',
            'Direct DPL Code':'023',
            'Direct DPL Invert':'False',
        }
    },

    # -- 8CALL90D Personality
    {
        'group_name': '8CALL90D Personality',
        'base_xpath': ".//Recset[@Name='Conventional Personality']//EmbeddedNode[@ReferenceKey='8CALL90D']",
        'context_node_name': 'Conventional Personality',
        'fields': {
            'Rx / TA Frequency (MHz)': '851.012500',
            'User Selectable PL (MPL)':'False',
            'Tx Squelch Type':'PL',
            'Tx DPL Code':'023',
            'Tx DPL Invert':'False',
            'Rx / TA Squelch Type':'PL',
            'Tx Frequency (MHz)':'851.012500',
            'Tx Network ID':'659',
            'Tx PL Code':'5A',
            'Tx PL Freq':'156.7',
            'Rx / TA  PL Code':'5A',
            'Rx / TA PL Freq':'156.7',
            'Rx / TA DPL Code':'023',
            'Rx / TA DPL Invert':'False',
            'Rx / TA  Network ID':'659',
            # 'Direct / Talkaround':'True',
            'Direct Squelch Type':'PL',
            'Direct PL Freq':'67.0',
            'Direct PL Code':'XZ',
            'Tx Deviation / Channel Spacing': '4 kHz / 20 kHz',
            'Name':'8CALL90D',
            'Direct Network ID':'659',
            'User Selectable PL [MPL]':'Disabled',
            'Direct Frequency (MHz)':'851.012500',
            'Direct DPL Code':'023',
            'Direct DPL Invert':'False',
        }
    },

    # -- 8TAC91 Personality
    {
        'group_name': '8TAC91 Personality',
        'base_xpath': ".//Recset[@Name='Conventional Personality']//EmbeddedNode[@ReferenceKey='8TAC91']",
        'context_node_name': 'Conventional Personality',
        'fields': {
            'Rx / TA Frequency (MHz)': '851.512500',
            'User Selectable PL (MPL)':'False',
            'Tx Squelch Type':'PL',
            'Tx DPL Code':'023',
            'Tx DPL Invert':'False',
            'Rx / TA Squelch Type':'PL',
            'Tx Frequency (MHz)':'806.512500',
            'Tx Network ID':'659',
            'Tx PL Code':'5A',
            'Tx PL Freq':'156.7',
            'Rx / TA  PL Code':'5A',
            'Rx / TA PL Freq':'156.7',
            'Rx / TA DPL Code':'023',
            'Rx / TA DPL Invert':'False',
            'Rx / TA  Network ID':'659',
            'Direct / Talkaround':'False',
            'Direct Squelch Type':'PL',
            'Direct PL Freq':'67.0',
            'Direct PL Code':'XZ',
            'Tx Deviation / Channel Spacing': '4 kHz / 20 kHz',
            'Name':'8TAC91',
            'Direct Network ID':'659',
            'User Selectable PL [MPL]':'Disabled',
            'Direct Frequency (MHz)':'851.512500',
            'Direct DPL Code':'023',
            'Direct DPL Invert':'False',
        }
    },

    # -- 8TAC91D Personality
    {
        'group_name': '8TAC91D Personality',
        'base_xpath': ".//Recset[@Name='Conventional Personality']//EmbeddedNode[@ReferenceKey='8TAC91D']",
        'context_node_name': 'Conventional Personality',
        'fields': {
            'Rx / TA Frequency (MHz)': '851.512500',
            'User Selectable PL (MPL)':'False',
            'Tx Squelch Type':'PL',
            'Tx DPL Code':'023',
            'Tx DPL Invert':'False',
            'Rx / TA Squelch Type':'PL',
            'Tx Frequency (MHz)':'851.512500',
            'Tx Network ID':'659',
            'Tx PL Code':'5A',
            'Tx PL Freq':'156.7',
            'Rx / TA  PL Code':'5A',
            'Rx / TA PL Freq':'156.7',
            'Rx / TA DPL Code':'023',
            'Rx / TA DPL Invert':'False',
            'Rx / TA  Network ID':'659',
            'Direct Squelch Type':'PL',
            'Direct PL Freq':'67.0',
            'Direct PL Code':'XZ',
            'Tx Deviation / Channel Spacing': '4 kHz / 20 kHz',
            'Name':'8TAC91D',
            'Direct Network ID':'659',
            'User Selectable PL [MPL]':'Disabled',
            'Direct Frequency (MHz)':'851.512500',
            'Direct DPL Code':'023',
            'Direct DPL Invert':'False',
        }
    },
    
    # -- 8TAC92 Personality --
    {
        'group_name': '8TAC92 Personality',
        'base_xpath': ".//Recset[@Name='Conventional Personality']//EmbeddedNode[@ReferenceKey='8TAC92']",
        'context_node_name': 'Conventional Personality',
        'fields': {
            'Rx / TA Frequency (MHz)': '852.012500',
            'User Selectable PL (MPL)':'False',
            'Tx Squelch Type':'PL',
            'Tx DPL Code':'023',
            'Tx DPL Invert':'False',
            'Rx / TA Squelch Type':'PL',
            'Tx Frequency (MHz)':'807.012500',
            'Tx Network ID':'659',
            'Tx PL Code':'5A',
            'Tx PL Freq':'156.7',
            'Rx / TA  PL Code':'5A',
            'Rx / TA PL Freq':'156.7',
            'Rx / TA DPL Code':'023',
            'Rx / TA DPL Invert':'False',
            'Rx / TA  Network ID':'659',
            'Direct / Talkaround':'False',
            'Direct Squelch Type':'PL',
            'Direct PL Freq':'67.0',
            'Direct PL Code':'XZ',
            'Tx Deviation / Channel Spacing': '4 kHz / 20 kHz',
            'Name':'8TAC92',
            'Direct Network ID':'659',
            'User Selectable PL [MPL]':'Disabled',
            'Direct Frequency (MHz)':'852.012500',
            'Direct DPL Code':'023',
            'Direct DPL Invert':'False',
        }
    },
    
    # -- 8TAC92D Personality --
    {
        'group_name': '8TAC92D Personality',
        'base_xpath': ".//Recset[@Name='Conventional Personality']//EmbeddedNode[@ReferenceKey='8TAC92D']",
        'context_node_name': 'Conventional Personality',
        'fields': {
            'Rx / TA Frequency (MHz)': '852.012500',
            'User Selectable PL (MPL)':'False',
            'Tx Squelch Type':'PL',
            'Tx DPL Code':'023',
            'Tx DPL Invert':'False',
            'Rx / TA Squelch Type':'PL',
            'Tx Frequency (MHz)':'852.012500',
            'Tx Network ID':'659',
            'Tx PL Code':'5A',
            'Tx PL Freq':'156.7',
            'Rx / TA  PL Code':'5A',
            'Rx / TA PL Freq':'156.7',
            'Rx / TA DPL Code':'023',
            'Rx / TA DPL Invert':'False',
            'Rx / TA  Network ID':'659',
            # 'Direct / Talkaround':'True',
            'Direct Squelch Type':'PL',
            'Direct PL Freq':'67.0',
            'Direct PL Code':'XZ',
            'Tx Deviation / Channel Spacing': '4 kHz / 20 kHz',
            'Name':'8TAC92D',
            'Direct Network ID':'659',
            'User Selectable PL [MPL]':'Disabled',
            'Direct Frequency (MHz)':'852.012500',
            'Direct DPL Code':'023',
            'Direct DPL Invert':'False',
        }
    },

    # Personality
    # -- 8TAC93 Personality
    {
        'group_name': '8TAC93 Personality',
        'base_xpath': ".//Recset[@Name='Conventional Personality']//EmbeddedNode[@ReferenceKey='8TAC93']",
        'context_node_name': 'Conventional Personality',
        'fields': {
            'Rx / TA Frequency (MHz)': '852.512500',
            'User Selectable PL (MPL)':'False',
            'Tx Squelch Type':'PL',
            'Tx DPL Code':'023',
            'Tx DPL Invert':'False',
            'Rx / TA Squelch Type':'PL',
            'Tx Frequency (MHz)':'807.512500',
            'Tx Network ID':'659',
            'Tx PL Code':'5A',
            'Tx PL Freq':'156.7',
            'Rx / TA  PL Code':'5A',
            'Rx / TA PL Freq':'156.7',
            'Rx / TA DPL Code':'023',
            'Rx / TA DPL Invert':'False',
            'Rx / TA  Network ID':'659',
            'Direct / Talkaround':'False',
            'Direct Squelch Type':'PL',
            'Direct PL Freq':'67.0',
            'Direct PL Code':'XZ',
            'Tx Deviation / Channel Spacing': '4 kHz / 20 kHz',
            'Name':'8TAC93',
            'Direct Network ID':'659',
            'User Selectable PL [MPL]':'Disabled',
            'Direct Frequency (MHz)':'852.512500',
            'Direct DPL Code':'023',
            'Direct DPL Invert':'False',
        }
    },

    # Personality
    # -- 8TAC93D
    {
        'group_name': '8TAC93D Personality',
        'base_xpath': ".//Recset[@Name='Conventional Personality']//EmbeddedNode[@ReferenceKey='8TAC93D']",
        'context_node_name': 'Conventional Personality',
        'fields': {
            'Rx / TA Frequency (MHz)': '852.512500',
            'User Selectable PL (MPL)':'False',
            'Tx Squelch Type':'PL',
            'Tx DPL Code':'023',
            'Tx DPL Invert':'False',
            'Rx / TA Squelch Type':'PL',
            'Tx Frequency (MHz)':'852.512500',
            'Tx Network ID':'659',
            'Tx PL Code':'5A',
            'Tx PL Freq':'156.7',
            'Rx / TA  PL Code':'5A',
            'Rx / TA PL Freq':'156.7',
            'Rx / TA DPL Code':'023',
            'Rx / TA DPL Invert':'False',
            'Rx / TA  Network ID':'659',
            # 'Direct / Talkaround':'True',
            'Direct Squelch Type':'PL',
            'Direct PL Freq':'67.0',
            'Direct PL Code':'XZ',
            'Tx Deviation / Channel Spacing': '4 kHz / 20 kHz',
            'Name':'8TAC93D',
            'Direct Network ID':'659',
            'User Selectable PL [MPL]':'Disabled',
            'Direct Frequency (MHz)':'852.512500',
            'Direct DPL Code':'023',
            'Direct DPL Invert':'False',
        }
    },
    # 
    # -- 8TAC94
    {
        'group_name': '8TAC94 Personality',
        'base_xpath': ".//Recset[@Name='Conventional Personality']//EmbeddedNode[@ReferenceKey='8TAC94']",
        'context_node_name': 'Conventional Personality',
        'fields': {
            'Rx / TA Frequency (MHz)': '853.012500',
            'User Selectable PL (MPL)':'False',
            'Tx Squelch Type':'PL',
            'Tx DPL Code':'023',
            'Tx DPL Invert':'False',
            'Rx / TA Squelch Type':'PL',
            'Tx Frequency (MHz)':'808.012500',
            'Tx Network ID':'659',
            'Tx PL Code':'5A',
            'Tx PL Freq':'156.7',
            'Rx / TA  PL Code':'5A',
            'Rx / TA PL Freq':'156.7',
            'Rx / TA DPL Code':'023',
            'Rx / TA DPL Invert':'False',
            'Rx / TA  Network ID':'659',
            'Direct / Talkaround':'False',
            'Direct Squelch Type':'PL',
            'Direct PL Freq':'67.0',
            'Direct PL Code':'XZ',
            'Tx Deviation / Channel Spacing': '4 kHz / 20 kHz',
            'Name':'8TAC94',
            'Direct Network ID':'659',
            'User Selectable PL [MPL]':'Disabled',
            'Direct Frequency (MHz)':'853.012500',
            'Direct DPL Code':'023',
            'Direct DPL Invert':'False',
        }
    },

    # -- 8TAC94D
    {
        'group_name': '8TAC94D Personality',
        'base_xpath': ".//Recset[@Name='Conventional Personality']//EmbeddedNode[@ReferenceKey='8TAC94D']",
        'context_node_name': 'Conventional Personality',
        'fields': {
            'Rx / TA Frequency (MHz)': '853.012500',
            'User Selectable PL (MPL)':'False',
            'Tx Squelch Type':'PL',
            'Tx DPL Code':'023',
            'Tx DPL Invert':'False',
            'Rx / TA Squelch Type':'PL',
            'Tx Frequency (MHz)':'853.012500',
            'Tx Network ID':'659',
            'Tx PL Code':'5A',
            'Tx PL Freq':'156.7',
            'Rx / TA  PL Code':'5A',
            'Rx / TA PL Freq':'156.7',
            'Rx / TA DPL Code':'023',
            'Rx / TA DPL Invert':'False',
            'Rx / TA  Network ID':'659',
            # 'Direct / Talkaround':'True',
            'Direct Squelch Type':'PL',
            'Direct PL Freq':'67.0',
            'Direct PL Code':'XZ',
            'Tx Deviation / Channel Spacing': '4 kHz / 20 kHz',
            'Name':'8TAC94D',
            'Direct Network ID':'659',
            'User Selectable PL [MPL]':'Disabled',
            'Direct Frequency (MHz)':'853.012500',
            'Direct DPL Code':'023',
            'Direct DPL Invert':'False',
        }
    },

    # ----------------------------------------------------------
    # ------ Channels ------
    # ----------------------------------------------------------

    # -- GW IO 1 --
    {
        'group_name': 'INTEROP - GW IO 1',
        'base_xpath': ".//Recset[@Name='Zone Channel Assignment']/Node[contains(@ReferenceKey, 'INTEROP')]//EmbeddedNode[@ReferenceKey='1-GW IO 1']",
        'context_node_name': 'Zone Channel Assignment',
        'fields': {
            'Channel Type': 'Trk',
            'Personality': '027A - IO',
            'Channel Name': 'GW IO 1',
            'Top Display Channel': 'GW IO 1',
            'Trunking Talkgroup': 'IO 1',
            'Active Channel': 'True'
        }
    },
    
    # -- GW IO 2 --
    {
        'group_name': 'INTEROP - GW IO 2',
        'base_xpath': ".//Recset[@Name='Zone Channel Assignment']/Node[contains(@ReferenceKey, 'INTEROP')]//EmbeddedNode[@ReferenceKey='2-GW IO 2']",
        'context_node_name': 'Zone Channel Assignment',
        'fields': {
            'Channel Type': 'Trk',
            'Personality': '027A - IO',
            'Channel Name': 'GW IO 2',
            'Top Display Channel': 'GW IO 2',
            'Trunking Talkgroup': 'IO 2',
            'Active Channel': 'True'
        }
    },
    
    # -- GW IO 3 --
    {
        'group_name': 'INTEROP - GW IO 3',
        'base_xpath': ".//Recset[@Name='Zone Channel Assignment']/Node[contains(@ReferenceKey, 'INTEROP')]//EmbeddedNode[@ReferenceKey='3-GW IO 3']",
        'context_node_name': 'Zone Channel Assignment',
        'fields': {
            'Channel Type': 'Trk',
            'Personality': '027A - IO',
            'Channel Name': 'GW IO 3',
            'Top Display Channel': 'GW IO 3',
            'Trunking Talkgroup': 'IO 3',
            'Active Channel': 'True'
        }
    },

    # -- GW IO 4 --
    {
        'group_name': 'INTEROP - GW IO 4',
        'base_xpath': ".//Recset[@Name='Zone Channel Assignment']/Node[contains(@ReferenceKey, 'INTEROP')]//EmbeddedNode[@ReferenceKey='4-GW IO 4']",
        'context_node_name': 'Zone Channel Assignment',
        'fields': {
            'Channel Type': 'Trk',
            'Personality': '027A - IO',
            'Channel Name': 'GW IO 4',
            'Top Display Channel': 'GW IO 4',
            'Trunking Talkgroup': 'IO 4',
            'Active Channel': 'True'
        }
    },

    # -- GW IO 5 --
    {
        'group_name': 'INTEROP - GW IO 5',
        'base_xpath': ".//Recset[@Name='Zone Channel Assignment']/Node[contains(@ReferenceKey, 'INTEROP')]//EmbeddedNode[@ReferenceKey='5-GW IO 5']",
        'context_node_name': 'Zone Channel Assignment',
        'fields': {
            'Channel Type': 'Trk',
            'Personality': '027A - IO',
            'Channel Name': 'GW IO 5',
            'Top Display Channel': 'GW IO 5',
            'Trunking Talkgroup': 'IO 5',
            'Active Channel': 'True'
        }
    },
 
    # -- GW IO 6 --
    {
        'group_name': 'INTEROP - GW IO 6',
        'base_xpath': ".//Recset[@Name='Zone Channel Assignment']/Node[contains(@ReferenceKey, 'INTEROP')]//EmbeddedNode[@ReferenceKey='6-GW IO 6']",
        'context_node_name': 'Zone Channel Assignment',
        'fields': {
            'Channel Type': 'Trk',
            'Personality': '027A - IO',
            'Channel Name': 'GW IO 6',
            'Top Display Channel': 'GW IO 6',
            'Trunking Talkgroup': 'IO 6',
            'Active Channel': 'True'
        }
    },

    # -- 8CALL90 Channel--
    {
        'group_name': 'INTEROP - 8CALL90',
        'base_xpath': ".//Recset[@Name='Zone Channel Assignment']/Node[contains(@ReferenceKey, 'INTEROP')]//EmbeddedNode[@ReferenceKey='7-8CALL90']",
        'context_node_name': 'Zone Channel Assignment',
        'fields': {
            'Channel Type': 'Cnv',
            'Personality': '8TAC',
            'Channel Name': '8CALL90',
            'Top Display Channel': '8CALL90',
            'Active Channel': 'True'
        }
    },

]

def _get_unit_id_for_system(root, system_name_contains):
    """
    Returns an integer ID for a Trunking System whose ReferenceKey contains the given name.
    """
    # for the translate function
    upper_abc = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'
    lower_abc = 'abcdefghijklmnopqrstuvwxyz'

    # Convert your search term to lowercase in Python
    search_term_lower = system_name_contains.lower()

    xpath = f".//Recset[@Name='Trunking System']/Node[contains(translate(@ReferenceKey, '{upper_abc}', '{lower_abc}'), '{search_term_lower}')]/Section[@Name='General']/Field[@Name='Unit ID']"
    
    elements = root.xpath(xpath)
    if elements and elements[0].text:
        try:
            return int(elements[0].text.strip())
        except (ValueError, TypeError):
            print(f"Warning: Could not convert Unit ID for '{system_name_contains}' to an integer.")
    # return nothing if empty or invalid

def _extract_metadata(root):
    metadata={
        "alias": "Unknown",
        "gwinnett_id": 0,
        "dekalb_id": 0,
        "hall_id": 0,
        "cobb_id": 0,
        "atlanta_id": 0,
        "fulton_id": 0
    }

    # Extract Alias
    alias_elements = root.xpath(".//Recset[@Name='Radio Wide']//Field[@Name='User Information\\Radio Alias']")
    if alias_elements and alias_elements[0].text:
        metadata["alias"] = alias_elements[0].text.strip()
    
    # Extract Unit IDs for each system
    metadata["gwinnett_id"] = _get_unit_id_for_system(root, "GWINNETT")
    metadata["dekalb_id"] = _get_unit_id_for_system(root, "Dekalb")
    metadata["hall_id"] = _get_unit_id_for_system(root, "Hall")
    metadata["cobb_id"] = _get_unit_id_for_system(root, "UASI")
    metadata["atlanta_id"] = _get_unit_id_for_system(root, "Atlanta")
    metadata["fulton_id"] = _get_unit_id_for_system(root, "FULTON")

    return metadata

# display problems
def _process_check_group(root, group, metadata, serial, model, mobile_hh):
    error_rows = []
    group_name = group['group_name']
    parents = root.xpath(group['base_xpath'])

    if not parents:
        error_rows.append([serial, metadata['alias'], metadata['gwinnett_id'], "N/A", group_name, "N/A", "Section Missing", "N/A", "N/A", model, mobile_hh, metadata['dekalb_id'], "dekalb", metadata['fulton_id'], "Fulton", metadata['atlanta_id'], "Atlanta", metadata['cobb_id'], "cobb", metadata['hall_id'], "hall"])
        return error_rows

    for parent in parents:
        system_context = "N/A"
        context_name = group.get('context_node_name') # Get the context to search for
        if context_name:
            context_xpath = f"ancestor::Node[@Name='{context_name}'][1]/@ReferenceKey"
            context_keys = parent.xpath(context_xpath)
            if context_keys:
                system_context = context_keys[0]

        for field_name, expected_value in group['fields'].items():
            if mobile_hh == 'Mobile' and field_name == 'Top Display Channel':
                continue # Skip 'Top Display Channel': field for Mobile or Console radios
            field_elements = parent.xpath(f".//Field[@Name='{field_name}']")

            if not field_elements:
                error_rows.append([serial, metadata['alias'], metadata['gwinnett_id'], system_context, group_name, field_name, "Setting Missing", expected_value, "N/A", model, mobile_hh, metadata['dekalb_id'], "dekalb", metadata['fulton_id'], "Fulton", metadata['atlanta_id'], "Atlanta", metadata['cobb_id'], "cobb", metadata['hall_id'], "hall"])
                continue

            actual_value = field_elements[0].text or ""
            if actual_value != expected_value:
                error_rows.append([serial, metadata['alias'], metadata['gwinnett_id'], system_context, group_name, field_name, "Incorrect Value", expected_value, actual_value, model, mobile_hh, metadata['dekalb_id'], "dekalb", metadata['fulton_id'], "Fulton", metadata['atlanta_id'], "Atlanta", metadata['cobb_id'], "cobb", metadata['hall_id'], "hall"])
                
    return error_rows

SERIAL_PREFIX_MAP = {
    '426': (4000, 'Portable'),
    '481': (6000, 'Portable'),
    '527': (6500, 'Mobile'),
    '579': (8000, 'Portable'),
    '652': (6500, 'Mobile'),
    '681': (8500, 'Mobile'),
    '755': (6500, 'Portable'),
    '756': (6000, 'Portable'),
    '761': (7500, 'Mobile'),
}

def _get_model_and_mobile_from_serial(serial):
    default = (0, 'Is Serial Correct?')
    prefix = serial[:3]
    return SERIAL_PREFIX_MAP.get(prefix, default)

def _get_model_from_filename(serial):
    serial_upper = serial.upper()
    if '4000' in serial_upper:
        return 4000
    elif '6000' in serial_upper:
        return 6000
    elif '6500' in serial_upper:
        return 6500
    elif '8000' in serial_upper:
        return 8000
    elif '8500' in serial_upper:
        return 8500
    elif '7500' in serial_upper:
        return 7500
    else:
        return 'Unknown'
    
def _get_mobile_from_filename(serial):
    serial_upper = serial.upper()
    if 'MOBILE' in serial_upper or 'MOB' in serial_upper or 'CONSOLE' in serial_upper or 'FSA' in serial_upper:
        return 'Mobile'
    elif 'HANDHELD' in serial_upper or 'HH' in serial_upper or 'PORTABLE' in serial_upper:
        return 'Portable'
    else:
        return _get_mobile_from_model(serial)

def _get_mobile_from_model(serial):
    model = _get_model_from_filename(serial)
    if model in [6500, 7500, 8500]:
        return 'Mobile'
    elif model in [4000, 6000]:
        return 'Portable'
    elif model == 8000:
        return 'Need description'
    else:
        return 'Is Type in Filename?'

def _validate_talkgroup_match(root, metadata, filename):
    """
    Any 'ASTRO Talkgroup ID' matches its corresponding 
    'Talkgroup Alias Text' and 'ReferenceKey'
    Returns a list of error rows if any mismatches are found.
    """
    error_rows = []    
    talkgroup_definitions = {} # 1. Build a map of all defined Talkgroup Aliases.
    definition_nodes = root.xpath(".//Recset[@Name='ASTRO Talkgroup List']//EmbeddedNode[@Name='Talkgroup Table']")

    for node in definition_nodes:
        ref_key = node.get('ReferenceKey') #  Key = ReferenceKey, Value = Alias Text.
        alias_text_elements = node.xpath(".//Field[@Name='Talkgroup Alias Text']")
        if ref_key and alias_text_elements and alias_text_elements[0].text is not None:
            talkgroup_definitions[ref_key] = alias_text_elements[0].text.strip()
    
    # 2. Check every 'ASTRO Talkgroup ID' field in the file.
    id_usage_fields = root.xpath(".//Field[@Name='ASTRO Talkgroup ID']")
    for field in id_usage_fields:
        used_id = field.text.strip() if field.text else ""
        
        if used_id in ["TG 1", ""]: # Ignore the default "TG 1" case
            continue

        # Does the used ID exist as a key, and does its value also match?
        if talkgroup_definitions.get(used_id) == used_id:
            continue  # This is the success case: all three strings match.

        # Something is wrong if we reach here
        context_node = field.xpath("ancestor::*[@ReferenceKey][1]")
        context_key = context_node[0].get('ReferenceKey') if context_node else "Unknown Context"

        if used_id not in talkgroup_definitions:
            issue = "Undeclared Talkgroup ID"
            expected = "A declared Talkgroup Alias"
            actual = used_id
        else:
            issue = "Inconsistent Definition"
            expected = f"Alias Text to match ReferenceKey ('{used_id}')"
            actual = talkgroup_definitions.get(used_id, "Not Found")
        
        error_rows.append([
            filename, metadata['alias'], metadata['gwinnett_id'], context_key, 
            "Talkgroup Consistency", f"ASTRO Talkgroup ID: {used_id}", 
            issue, expected, actual
        ])
    
    return error_rows

# Check XML file
def check_xml_file(filepath, report_rows):
    try:
        parser = ETREE.XMLParser(remove_blank_text=True, resolve_entities=False)
        tree = ETREE.parse(filepath, parser)
        root = tree.getroot()
        filename = os.path.basename(filepath)
        serial = filename.removesuffix('.xml')

        if len(serial)==10:
            model, mobile = _get_model_and_mobile_from_serial(serial)
        else:
            model = _get_model_from_filename(serial)
            mobile = _get_mobile_from_filename(serial)

        metadata = _extract_metadata(root)

        discrepancies_in_file = []
        
        for group in CHECKS_TO_PERFORM:
            errors = _process_check_group(root, group, metadata, serial, model, mobile)
            if errors:
                discrepancies_in_file.extend(errors)
        
        talkgroup_errors = _validate_talkgroup_match(root, metadata, serial)
        if talkgroup_errors:
            discrepancies_in_file.extend(talkgroup_errors)

        if not discrepancies_in_file:
            success_row = [serial, metadata['alias'], metadata['gwinnett_id'], "OK", "OK", "OK", "OK", "OK", "OK", model, mobile, metadata['dekalb_id'], "DEK", metadata['fulton_id'], "Fulton", metadata['atlanta_id'], "Atl", metadata['cobb_id'], "17D", metadata['hall_id'], "1DE" , "td-gw", "TD-Alias"]
            report_rows.append(success_row)
            return False
        else:
            report_rows.extend(discrepancies_in_file)
            return True

    except ETREE.XMLSyntaxError:
        # this should not happen due to prior validation
        print(f"Error: Could not parse XML file '{filepath}'.")
        report_rows.append([os.path.basename(filepath), "Error!", "Alias", "ID", "Setting", "Ref", "Group", "Could not parse XML", "Expect", "Actual", "model", "type", "Dekalb", "1F5","Fulton","5B2","Atlanta","293","Cobb", "UASI", "Hall", "TD-Hal", "TD-Gw", "TD-Alias"])
        return True

# Adjust Excel column widths
def adjust_column_width(worksheet):
    for col_cells in worksheet.columns:
        max_length = 0
        col_letter = col_cells[0].column_letter
        for cell in col_cells:
            if cell.value is not None:
                cell_len = len(str(cell.value))
                max_length = max(max_length, cell_len)
        
        try:
            header_row = worksheet[1]
            col_i = column_index_from_string(col_letter) - 1
            header_cell = header_row[col_i]
            header_len = len(str(header_cell.value)) # header length if longer
            max_length = max(max_length, header_len)
        except(IndexError, TypeError):
            pass

        worksheet.column_dimensions[col_letter].width = max_length + 1.5 #Return padding

# Generate Excel report
def _generate_report(report_filename, report_rows, files_with_errors, total_files):
    with pd.ExcelWriter(report_filename, engine='openpyxl') as writer:
    
        header = ['Filename', 'Alias', 'Gw ID', 'Setting','Reference', 'Group','Problem', 'Expected', 'Actual', 'Model', 'Type', 'Dekalb', 'TD-Dek', 'Fulton', 'TD-Ful', 'Atlanta', 'TD-Atl', 'Cobb', 'TD-Cob', 'Hall', 'TD-Hal', 'TD-Gw', 'TD-Alias']
        df = pd.DataFrame(report_rows, columns=header)
        sheet_name = f'{files_with_errors} of {total_files} files have errors'
        df.to_excel(writer, sheet_name=sheet_name, index=False)

        worksheet = writer.sheets[sheet_name]

        BLACK = '00000000'
        WHITE = '00FFFFFF'
        BLUE = '000000FF'
        GREEN = '38761D'
        RED = '990000'
        GRAY = '00C0C0C0'
    
        black_fill = PatternFill(start_color=BLACK, end_color=BLACK, fill_type="solid") # Black fill
        green_fill = PatternFill(start_color=GREEN, end_color=GREEN, fill_type="solid") # Green fill
        red_fill = PatternFill(start_color=RED, end_color=RED, fill_type="solid") # Red fill

        for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, min_col=1, max_col=16):
            for cell in row:
                cell.fill = black_fill
                cell.font = Font(bold=False, size=11, color=WHITE, name='Arial') # White font
                cell.alignment = Alignment(horizontal='left', vertical='center')

        # Header
        for cell in worksheet[1]:
            cell.font = Font(bold=True, size=12, color=WHITE) # White font
            cell.fill = PatternFill(start_color=BLUE, end_color=BLUE, fill_type="solid") # Blue fill
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # Data rows with 22 columns
        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=1, max_col=len(header)):

            dekalb_id = row[11]  # Dekalb TD column (12th column, index 11)
            dekalb_td_id = row[12] # Dekalb TD column (13th column, index 12)
            fulton_id = row[13]  # Fulton ID column (14th column, index 13)
            fulton_td_id = row[14] # Fulton TD column (15th column, index 14)
            atlanta_id = row[15]  # Atlanta ID column (16th column,
            atlanta_td_id = row[16] # Atlanta TD column (17th column, index 16)
            cobb_id = row[17]   # Cobb ID column (18th column, index 17)
            cobb_td_id = row[18] # Cobb TD column (19th column, index 18)
            hall_id = row[19]   # Hall ID column (20th column, index 19)
            hall_td_id = row[20] # Hall TD column (21th column, index
            gwinnett_id = row[2]  # Gwinnett ID column (3rd column, index 2)
            gwinnett_td_id = row[21] # Gwinnett TD column (22th column, index 21)
            alias = row[1]  # Alias column (2nd column, index 1)
            td_alias = row[22] # TD-Gw column (23th column, index 22)

            for cell in row:
                cell.fill = black_fill
                cell.font = Font(bold=False, size=11, color=WHITE) # White font for data
                cell.alignment = Alignment(horizontal='left', vertical='center')
                cell.border = Border(
                    left=Side(border_style="thin", color=GRAY),
                    right=Side(border_style="thin", color=GRAY),
                    top=Side(border_style="thin", color=GRAY),
                    bottom=Side(border_style="thin", color=GRAY)
                )

                if cell.value == "OK":
                    cell.fill = green_fill

                if cell.column == 7: # Problem = G (7th column)
                    if cell.value == "Section Missing":
                        cell.fill = red_fill

                if cell.column == 9: # Actual Problem = I (9th column)
                    if cell.value != "OK":
                        cell.fill = red_fill

                if cell.column == 12 and dekalb_id.value == dekalb_td_id.value:
                    cell.fill = green_fill
                elif cell.column == 12 and dekalb_id.value != dekalb_td_id.value:
                    cell.fill = red_fill

                if cell.column == 14 and fulton_id.value == fulton_td_id.value:
                    cell.fill = green_fill
                elif cell.column == 14 and fulton_id.value != fulton_td_id.value:
                    cell.fill = red_fill

                if cell.column == 16 and atlanta_id.value == atlanta_td_id.value:
                    cell.fill = green_fill
                elif cell.column == 16 and atlanta_id.value != atlanta_td_id.value:
                    cell.fill = red_fill

                if cell.column == 18 and cobb_id.value == cobb_td_id.value:
                    cell.fill = green_fill
                elif cell.column == 18 and cobb_id.value != cobb_td_id.value:
                    cell.fill = red_fill

                if cell.column == 20 and hall_id.value == hall_td_id.value:
                    cell.fill = green_fill
                elif cell.column == 20 and hall_id.value != hall_td_id.value:
                    cell.fill = red_fill

                if cell.column == 3 and gwinnett_id.value == gwinnett_td_id.value:
                    cell.fill = green_fill
                elif cell.column == 3 and gwinnett_id.value != gwinnett_td_id.value:
                    cell.fill = red_fill
                
                if cell.column == 2 and alias.value == td_alias.value:
                    cell.fill = green_fill
                elif cell.column == 2 and alias.value != td_alias.value:
                    cell.fill = red_fill

        worksheet.freeze_panes = "B2" # Freeze top row & first column

        adjust_column_width(worksheet)

    print(f"Opening Report: {report_filename}")
    try:
        os.startfile(report_filename) # open the report
    except AttributeError:
        print("Open report manually.")


###### Main function ######

def main():

    print("Motorola Codeplug Checker")
    print("by Morgan King, Gwinnett County")
    use_api = True
    df_td = None # DataFrame for Radio assets from TD

    # Fetch Radio Assets from API
    print("Fetching Radio Assets from API...")
    # --- Configuration for Sandbox Environment ---
    BASE_URL = "https://support.gwinnettcounty.com/SBTDWebApi"
    ASSET_APP_ID = 279
    RADIO_ASSET_FORM_ID = 9856
    USERNAME = os.getenv("TD_USERNAME")
    PASSWORD = os.getenv("TD_PASSWORD")

    client = TeamDynamixSandboxClient(base_url=BASE_URL, asset_app_id=ASSET_APP_ID)

    if client.authenticate(username=USERNAME, password=PASSWORD):
        radio_assets = client.get_all_assets(form_id=RADIO_ASSET_FORM_ID)

        if radio_assets:
            df_td = pd.DataFrame(radio_assets)
            print(f"Fetched {len(df_td)} radio assets from API.")
            df_td.to_excel('RadioAssets.xlsx', index=False)
            print("Saved radio assets to 'RadioAssets.xlsx'.")
            print(df_td.head()) # Display first few rows
        else:
            use_api = False
            print("Failed to fetch radio assets from API.")
    else:
        use_api = False
        print("Failed to authenticate with API.")

    # Load TD.xlsx if API fetch not used or failed
    if not use_api:
        print("Loading TD.xlsx...")
        td_file = 'TD.xlsx'
        
        try:
            df_td = pd.read_excel(td_file)
            print("TD.xlsx loaded.")
        except FileNotFoundError:
            print(f"Warning: '{td_file}' not found in the current directory.")
        except ImportError:
            print("Error: 'openpyxl' library is required to read Excel files.")
            print("Install it using 'pip install openpyxl'")
            input("Press Enter to exit...") # hold terminal open
            return    

    print("Checking XML Codeplugs...")
    xml_files = glob.glob('*.xml') # Find all XML files in folder
    if not xml_files:
        print("No XML Codeplugs in this folder.")
        print("Run this program in a folder with XML Codeplugs.")
        input("Press Enter to exit...") # hold terminal open
        return

    total_files = len(xml_files)
    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M")
    report_filename = f'Codeplug-Report_{timestamp}.xlsx'
    report_rows = []
    files_with_errors = 0

    # input each row
    for i, filepath in enumerate(xml_files):
        print(f"Processing file {i+1} of {total_files}: {os.path.basename(filepath)}")
        if check_xml_file(filepath, report_rows):
            files_with_errors += 1

    # add data from TD.xlsx to report
    rows_with_td = []
    if df_td is not None:
        serial_col = 'SerialNumber' if use_api else 'Serial Number'
        for row in report_rows:
            serial = str(row[0]).strip()
            td_match = df_td[df_td[serial_col] == serial]
            if not td_match.empty:
                td_dekalb = td_match.iloc[0].get('(1F5) Dekalb', 'N/A')
                td_fulton = td_match.iloc[0].get('(5B2) Fulton', 'N/A')
                td_atlanta = td_match.iloc[0].get('(293) Atlanta', 'N/A')
                td_cobb = td_match.iloc[0].get('(17D) Cobb', 'N/A')
                td_hall = td_match.iloc[0].get('(1DE) Hall', 'N/A')
                td_gwinnett = td_match.iloc[0].get('(027A) Gwinnett', 'N/A')
                td_alias = td_match.iloc[0].get('Radio User Alias', 'N/A')
                new_row = row[:12] + [td_dekalb, row[13], td_fulton, row[15], td_atlanta, row[17], td_cobb, row[19], td_hall, td_gwinnett, td_alias]
                rows_with_td.append(new_row)
            else:
                new_row = row + ['no Dekalb', row[13], 'no Fulton', row[15], 'no Atlanta', row[17], 'no Cobb', row[19], 'no Hall', 'no Gwinnett', 'no Alias']
                rows_with_td.append(new_row)
        report_rows = rows_with_td

    # Generate report
    _generate_report(report_filename, report_rows, files_with_errors, total_files)

if __name__ == "__main__":
    main()